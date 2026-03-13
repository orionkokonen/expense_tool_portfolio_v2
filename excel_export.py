# -*- coding: utf-8 -*-
"""
excel_export.py — Excel レポートの生成
シート: Errors / Warnings / Clean / Summary / Charts
見栄え: ヘッダ太字 / 先頭行固定 / オートフィルタ / 列幅自動調整
グラフ: 月別推移（棒グラフ）/ カテゴリ比率（円グラフ）
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
# Worksheet の型を import しておくと、関数の引数に「ws: Worksheet」と書ける。
# 型ヒント（型注釈）= 「この引数にはこの型が来る」と明示する仕組み。
# mypy（型チェックツール）がミスを見つけやすくなり、エディタの補完も効くようになる。
from openpyxl.worksheet.worksheet import Worksheet

from expense_core import SummaryRow


def write_xlsx_report(
    *,
    path: Path,
    errors: Sequence[Mapping[str, object]],
    warnings: Sequence[Mapping[str, object]],
    clean: Sequence[Mapping[str, object]],
    summary: Sequence[SummaryRow],
) -> None:
    """Excel レポートを生成して指定パスに保存する。

    シート構成:
      - Errors   : 入力チェックでエラーになった行
      - Warnings : 社内ルール違反の行
      - Clean    : 問題なしのクリーン行
      - Summary  : 集計結果
      - Charts   : グラフ（月別・カテゴリ別）

    出力先フォルダが存在しない場合は自動で作成する。
    Render のようなクラウド環境ではデプロイ時にフォルダがない場合があるため、
    mkdir を明示的に呼んでいる。
    """
    wb = Workbook()

    # Workbook 作成時に自動生成される空のシートを削除する。
    # 削除しないと意図しない空シートが先頭に残ってしまう。
    default = wb.active
    wb.remove(default)

    _add_table_sheet(
        wb, "Errors", errors, ["row", "date", "amount", "merchant", "category", "reason"]
    )
    _add_table_sheet(
        wb,
        "Warnings",
        warnings,
        ["kind", "row", "date", "month", "category", "merchant", "amount", "message"],
    )
    _add_table_sheet(wb, "Clean", clean, ["date", "amount", "merchant", "category"])
    _add_table_sheet(wb, "Summary", summary, ["type", "key", "value"])

    charts_ws = wb.create_sheet("Charts")
    _build_charts(charts_ws, summary)

    # 出力先の親フォルダがまだなくても、その場で作って保存できるようにする。
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _add_table_sheet(
    wb: Workbook, title: str, rows: Sequence[Mapping[str, object]], columns: Sequence[str]
) -> None:
    """データをテーブル形式でシートに書き込む。

    ヘッダを太字・先頭行固定・オートフィルタを設定して操作しやすくし、
    列幅は内容に合わせて自動調整する（_auto_width）。
    """
    ws = wb.create_sheet(title)

    # ヘッダ行を書き込み、太字・中央揃えにする
    ws.append(columns)
    header_font = Font(bold=True)
    for col_idx in range(1, len(columns) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    # データ行を書き込む
    for r in rows:
        ws.append([r.get(c, "") for c in columns])

    # 先頭行を固定し、フィルタを有効にする（データ量が多い場合の操作性向上）
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, ws.max_row)}"

    _auto_width(ws, max_width=50)


def _auto_width(ws: Worksheet, max_width: int = 60) -> None:
    """各列の最大文字数に応じて列幅を自動調整する。

    openpyxl はデフォルトで列幅を調整しないため、手動で計算して設定する。
    max_width で上限を設けることで、長い文字列で列が広がりすぎるのを防ぐ。
    """
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _build_charts(ws: Worksheet, summary: Sequence[SummaryRow]) -> None:
    """Charts シートに月別棒グラフとカテゴリ別円グラフを作る。

    summary のフラットリストから type でフィルタしてグラフ用データを抽出し、
    シートにテーブルとして書き込んでから openpyxl の Reference でグラフを作成する。
    テーブルをシートに残すことで Excel 上でもデータを確認・編集しやすくなる。
    """
    ws["A1"] = "Month totals"
    ws["A1"].font = Font(bold=True)

    # summary はフラット構造なので type でフィルタしてグラフ用データを取り出す
    month_rows = [
        (r["key"], int(r["value"]))
        for r in summary
        if r["type"] == "month_total" and r["key"] != "month"
    ]
    cat_rows = [
        (r["key"], int(r["value"]))
        for r in summary
        if r["type"] == "category_total" and r["key"] != "category"
    ]

    # 月別テーブルをシートに書き込む（A3 から）
    ws.append([])  # A2 を空行にする
    ws.append(["month", "total_amount"])  # row3: ヘッダ
    for m, v in month_rows:
        ws.append([m, v])

    month_start = 3
    month_end = 3 + len(month_rows)

    # 棒グラフ: データ範囲と軸ラベルを Reference で指定する
    if month_rows:
        chart = BarChart()
        chart.title = "Monthly total"
        chart.y_axis.title = "amount"
        chart.x_axis.title = "month"
        data = Reference(ws, min_col=2, min_row=month_start, max_row=month_end)
        cats = Reference(ws, min_col=1, min_row=month_start + 1, max_row=month_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        # 値ラベルを全部出すと棒が密集したときに読みにくくなるので、ここでは隠す。
        chart.dataLabels.showVal = False
        ws.add_chart(chart, "D3")

    # カテゴリテーブルを月別テーブルの直下に配置する
    row0 = month_end + 3
    ws[f"A{row0}"] = "Category totals"
    ws[f"A{row0}"].font = Font(bold=True)

    ws.append([])  # 空行
    ws.append(["category", "total_amount"])
    for c, v in cat_rows:
        ws.append([c, v])

    cat_header_row = row0 + 2
    cat_start = cat_header_row
    cat_end = cat_header_row + len(cat_rows)

    # 円グラフ: カテゴリ比率を可視化する
    if cat_rows:
        pie = PieChart()
        pie.title = "Category ratio"
        data = Reference(ws, min_col=2, min_row=cat_start, max_row=cat_end)
        labels = Reference(ws, min_col=1, min_row=cat_start + 1, max_row=cat_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws.add_chart(pie, f"D{cat_start}")
