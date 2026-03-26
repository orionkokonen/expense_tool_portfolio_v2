# このファイルは、Excel書き出し処理を確かめるテストです。
# -*- coding: utf-8 -*-
"""
tests/test_excel_export.py: excel_export.py の単体テスト

このファイルでは、「Excel ファイルが作れるか」だけでなく、
作られた中身が最低限正しいかも確認する。
見た目の細かい装飾を全部テストするのではなく、
壊れると困るポイントに絞るのが読みやすいテストのコツ。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_export import write_xlsx_report
from expense_core import ExpenseRowNorm, IssueRow, SummaryRow, WarningRow


def _sample_errors() -> list[IssueRow]:
    """Errors シート用の最小データ。

    テストでは「本物そっくりの大量データ」より、
    意図がすぐ分かる小さいデータのほうが読みやすい。
    """
    return [
        {
            "row": "2",
            "date": "2026/01/10",
            "amount": "abc",
            "merchant": "A",
            "category": "交通費",
            "reason": "日付形式が違う / 金額が数字ではない",
        }
    ]


def _sample_warnings() -> list[WarningRow]:
    """Warnings シート用の最小データ。"""
    return [
        {
            "kind": "category_unknown",
            "row": "3",
            "date": "2026-01-10",
            "month": "2026-01",
            "category": "雑費",
            "merchant": "レストラン",
            "amount": "2000",
            "message": "未登録カテゴリ: 雑費",
        }
    ]


def _sample_clean() -> list[ExpenseRowNorm]:
    """Clean シート用の正常データ。

    `ExpenseRowNorm` は「正規化後の行」なので、amount は文字列ではなく整数。
    """
    return [
        {
            "row": "4",
            "date": "2026-01-10",
            "amount": 1200,
            "merchant": "カフェ",
            "category": "会議費",
        },
        {
            "row": "5",
            "date": "2026-02-15",
            "amount": 3500,
            "merchant": "ホテル",
            "category": "旅費",
        },
    ]


def _sample_summary() -> list[SummaryRow]:
    """Summary / Charts シート用の最小データ。

    グラフ生成は summary の中身に依存するので、
    Excel 出力テストでも summary を少しだけ用意しておく。
    """
    return [
        {"type": "month_total", "key": "month", "value": "total_amount"},
        {"type": "month_total", "key": "2026-01", "value": "1200"},
        {"type": "month_total", "key": "2026-02", "value": "3500"},
        {"type": "category_total", "key": "category", "value": "total_amount"},
        {"type": "category_total", "key": "会議費", "value": "1200"},
        {"type": "category_total", "key": "旅費", "value": "3500"},
        {"type": "stats", "key": "count", "value": "2"},
        {"type": "stats", "key": "average", "value": "2350"},
    ]


class TestWriteXlsxReport:
    """write_xlsx_report() の振る舞いを確認するテスト群。"""

    def test_creates_file(self, tmp_path: Path) -> None:
        """関数を呼ぶと .xlsx ファイルが実際に作られるか。"""
        xlsx = tmp_path / "report.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=_sample_errors(),
            warnings=_sample_warnings(),
            clean=_sample_clean(),
            summary=_sample_summary(),
        )
        assert xlsx.exists()

    def test_sheet_names(self, tmp_path: Path) -> None:
        """必要なシート名がそろっているか。

        ここが崩れると、出力仕様が変わったことにすぐ気づける。
        """
        xlsx = tmp_path / "report.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=_sample_errors(),
            warnings=_sample_warnings(),
            clean=_sample_clean(),
            summary=_sample_summary(),
        )
        wb = load_workbook(xlsx)
        assert set(wb.sheetnames) == {
            "Errors",
            "Warnings",
            "Clean",
            "Summary",
            "Charts",
        }

    def test_errors_sheet_has_header_and_row(self, tmp_path: Path) -> None:
        """Errors シートにヘッダと 1 行分のデータが入るか。"""
        xlsx = tmp_path / "report.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=_sample_errors(),
            warnings=_sample_warnings(),
            clean=_sample_clean(),
            summary=_sample_summary(),
        )
        ws = load_workbook(xlsx)["Errors"]
        assert ws.max_row == 2
        assert ws.cell(row=1, column=1).value == "row"

    def test_clean_sheet_row_count(self, tmp_path: Path) -> None:
        """Clean シートの行数が、ヘッダ 1 行 + データ件数になるか。"""
        xlsx = tmp_path / "report.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=[],
            warnings=[],
            clean=_sample_clean(),
            summary=_sample_summary(),
        )
        ws = load_workbook(xlsx)["Clean"]
        assert ws.max_row == 3

    def test_header_is_bold(self, tmp_path: Path) -> None:
        """ヘッダが太字か。

        見た目の全部は追わないが、最低限の読みやすさが保たれているかは確認する。
        """
        xlsx = tmp_path / "report.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=[],
            warnings=[],
            clean=_sample_clean(),
            summary=_sample_summary(),
        )
        ws = load_workbook(xlsx)["Clean"]
        assert ws.cell(row=1, column=1).font.bold is True

    def test_creates_parent_directory(self, tmp_path: Path) -> None:
        """親フォルダがなくても保存できるか。

        利用者は毎回フォルダを手で作るとは限らないので、
        出力側で自動作成できることを守る。
        """
        xlsx = tmp_path / "nested" / "dir" / "report.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=[],
            warnings=[],
            clean=[],
            summary=_sample_summary(),
        )
        assert xlsx.exists()

    def test_empty_data_still_creates_charts_sheet(self, tmp_path: Path) -> None:
        """データ 0 件でもファイル全体の形が壊れないか。"""
        xlsx = tmp_path / "empty.xlsx"
        write_xlsx_report(
            path=xlsx,
            errors=[],
            warnings=[],
            clean=[],
            summary=[],
        )
        wb = load_workbook(xlsx)
        assert "Charts" in wb.sheetnames
